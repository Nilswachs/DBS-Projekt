# fuer dieses Programm muessen openpyxl und psycopg2 installiert sein
# wurde getestet mit Python 2.7
import re
import psycopg2
from openpyxl import load_workbook

wb = load_workbook(filename='american-election-tweets.xlsx', read_only= True) # importiert die Exceldatei
ws = wb['american-election-tweets']
conn = psycopg2.connect(database="Election", user="postgres", password="postgres", host="localhost") # stellt Verbindung zur Datenbank her
cur = conn.cursor()

# erstellt Tabelle tweets
for row in range(2, ws.max_row + 1): # geht alle Zeilen durch
	for column in "ABCDEHI": # geht die genannten Spalten durch
		cell_name = "{}{}".format(column,row)
		ak_cell = ws[cell_name].value # Wert der aktuellen Zelle
		if column == "A": # weist den Spalteneintraegen einer Zeile Variablen zu
			a = ak_cell
		elif column == "B":
			b = ak_cell
		elif column == "C":
			c = ak_cell
		elif column == "D":
			d = ak_cell
		elif column == "E":
			e = ak_cell
		elif column == "H":
			h = ak_cell
		elif column == "I":
			i = ak_cell
	ak_time = e[:10] + " " + e[11:] # wandelt die Zeitangabe in einen gueltigen Timestamp um
	cur.execute("INSERT INTO tweets (id, text, handle, time, retweet_count, favorite_count, original_author) VALUES (%s, %s, %s, %s, %s, %s, %s)", (row-1, b, a, ak_time, h, i, d)) # erzeugt eine neue Zeile in der Tabelle tweets der Datenbank Elections
	conn.commit() # macht Aenderungen an der Datenbank permanent


hashtags = [] # alle Hashtags die Vorkommen (keine Duplikate)
hashtags2 = [] # alle Hashtags mit der Zeile, in der sie Vorkommen (#name, row) (Hashtags koennen mehrfach vorkommen)
for row in range(2, ws.max_row+1): # geht durch alle Zeilen
	for column in "B": # geht durch Spalte B
		cell_name = "{}{}".format(column,row)
		text = ws[cell_name].value # text in der aktuellen Zelle
		hashtag = "" # hier wird ein gefundener Hashtag gespeichert
		i = 0 # Zaehlvariable
		while i < len(text):
			if text[i] == "#": # wenn ein Hashtag gefunden wird, werden solange alle folgenden Chars an hashtag angehaengt, bis ein "Abbruchszeichen gefunden wird (., !, ,, ?, ", usw.) 
				hashtag += text[i]
				i = i+1
				while text[i] != " " and text[i] != "!" and text[i] != "." and text[i] != "?" and text[i] != '\n' and text[i] != "," and text[i] != "#":
					hashtag += text[i]
					i = i+1
					if i == len(text): # verhindert dass i out of range ist
						break
			if hashtag != "": # wenn ein Hashtag gefunden wird, werden folgende Dinge gemacht:
				hashtag = ("".join(re.findall(r"[A-Za-z0-9#']*", hashtag))).replace(" ","") # entfernt alle Sonderzeichen entfernt ausser: # und ' (notwendig, da gelengtlich " am Anfang oder Ende des Hashtags auftritt
				if hashtag not in hashtags: # wenn der Hashtag noch nicht vorhanden ist, wird er eingefuegt
					hashtags.append(hashtag) 
				hashtags2 += [(hashtag, row)] 
				hashtag = "" # hashtag wird zurueckgesetzt
			if i >= len(text)-1:
				break
			if text[i] != "#":
				i = i+1
# erstellte Tabelle hashtags:
for i in range(0, len(hashtags)): # fuegt alle Hashtags mit ihrer ID in die Tabelle hashtags ein
	cur.execute("INSERT INTO hashtags (id, name) VALUES (%s, %s)", (i+1, hashtags[i]))
	conn.commit()

# erstellt Tabelle enthalten:
for i in range(0, len(hashtags2)): # erstellt die enthalten-Tabelle
	k = 0
	for e in range(0, len(hashtags2)):
		if i != e and hashtags2[i][1] == hashtags2[e][1]: # wenn die Zeile gleich ist werden die entsprechenden HashtagIDs und die TweetID in die Tabelle eingefuegt
			k = 1
			i_index = hashtags.index(hashtags2[i][0]) + 1
			e_index = hashtags.index(hashtags2[e][0]) + 1
			cur.execute("INSERT INTO enthalten (tweetid, hashtagid, mentioned_with) VALUES (%s, %s, %s)", (hashtags2[i][1]-1, i_index, e_index))
			conn.commit()
	if k==0: # wenn der Hastag alleine in einem Tweet vorkommt, wird er an dieser Stelle in die Tabelle eingefuegt
		cur.execute("INSERT INTO enthalten (tweetid, hashtagid) VALUES (%s, %s)", (hashtags2[i][1]-1, hashtags.index(hashtags2[i][0])+1))
		conn.commit()


