# fuer dieses Programm muss openpyxl installiert sein
# wurde getestet mit Python 2.7
from openpyxl import load_workbook

wb = load_workbook(filename='american-election-tweets.xlsx', read_only= True) # importiert die Exceldatei
ws = wb['american-election-tweets']
print(ws.calculate_dimension()) # gibt die Dimensionen der Tabelle an

for row in range(2, ws.max_row): # geht alle Zeilen durch
	for column in "ABCDEHI": # geht der Reihe nach die genannten Spalten durch
		cell_name = "{}{}".format(column,row) # der aktuelle Zellenindex (column, row)
		ak_cell = ws[cell_name].value # der Wert der aktuellen Zelle
		if column == "A": # gibt den Zellen einer Zeile verschiedene Variablen zum leichteren Zugriff
			a = ak_cell
		elif column == "B":
			b = ak_cell
		elif column == "C":
			c = ak_cell
		elif column == "D":
			d = ak_cell
		elif column == "E":
			e = ak_cell
		elif column == "H":
			h = ak_cell
		elif column == "I":
			i = ak_cell
	if a != "HillaryClinton" and a != "realDonaldTrump": # ueberprueft, ob der Post entweder von Hillary Clinton oder Donald Trump ist
		print("Fehler in Zeile " + str(row) + ", Spalte A")
	if b == "": # ueberprueft, ob es einen Text gibt
		print("Fehler in Zeile " + str(row) + ", Spalte B")
	if c == "True" and d == "": # ueberprueft, ob is_retweet mit original_author uebereinstimmt
		print("Fehler in Zeile " + str(row) + ", Spalte C oder D")
	elif c == "False" and d is not None: 
		print("Fehler in Zeile " + str(row) + ", Spalte C oder D")
	if e[:4] != "2016" or e[5:7] < "01" or e[5:7] > "12" or e[8:10] < "01" or e[8:10] > "31" or e[11:13] < "00" or e[11:13] > "23" or e[14:16] < "00" or e[14:16] > "59" or e[17:] < "00" or e[17:] > "59": # ueberprueft, ob es sich um einen gueltigen Timestamp handelt
		print("Fehler in Zeile " + str(row) + ", Spalte E")
	if isinstance(h,(int,long)) == False: # ueberprueft, ob retweet_count ein int ist
		print("Fehler in Zeile " + str(row) + ", Spalte H")
	if isinstance(i,(int,long)) == False: # ueberprueft, ob favorite_count ein int ist
		print("Fehler in Zeile " + str(row) + ", Spalte I")

# Diese Programm ist bei uns ohne Ausgabe durchgelaufen, d.h es wurden keine Fehler gefunden.
# Wir haben dann die relevanten Spalten des Datensatzes in unsere Datenbank importiert.
# siehe Aufgabe 3



